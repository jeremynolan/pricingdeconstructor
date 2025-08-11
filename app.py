from flask import Flask, request, make_response, session
import pandas as pd
import plotly.express as px
import plotly.io as pio
import os
import logging
from datetime import datetime
import openpyxl
import re
import string
import secrets

app = Flask(__name__)
app.secret_key = secrets.token_hex(16)  # Generate a secure random secret key
app.config['SESSION_COOKIE_HTTPONLY'] = True  # Enhance session security
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # Prevent CSRF
app.config['PERMANENT_SESSION_LIFETIME'] = 3600  # Session persists for 1 hour
UPLOAD_FOLDER = 'Uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Set up logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Store pricing rules in memory
pricing_rules = {
    "Process": {},
    "Coating": {},
    "Foil Material": {},
    "Foil Thickness": {},
    "Colour": {}
}

# Process and Step Process coupling
process_step_mapping = {
    "Chemetch": ["Single", "Double", "Triple", "5 or more"],
    "LaserSTEP": ["1-2", "1-5", "1-10", "1-15", "1-20", "21-30", "31-40", "41-50", "51-60"],
    "Milled": ["Single", "Double", "Triple", "Quad"],
    "LaserCut": []
}

# Inline CSS
css = """
<style>
    body { font-family: Arial, sans-serif; margin: 20px; background-color: #f4f4f9; }
    .container { max-width: 1200px; margin: 0 auto; }
    h1 { color: #333; text-align: center; }
    .form-group { margin-bottom: 15px; }
    label { display: inline-block; width: 200px; font-weight: bold; }
    input[type="file"], input[type="number"] { padding: 5px; width: 200px; }
    button { padding: 10px 20px; background-color: #007bff; color: white; border: none; cursor: pointer; }
    button:hover { background-color: #0056b3; }
    table { width: 100%; border-collapse: collapse; margin-top: 20px; }
    th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
    th { background-color: #007bff; color: white; }
    tr:nth-child(even) { background-color: #f2f2f2; }
    .error { color: red; text-align: center; }
    .debug { color: blue; text-align: left; }
    .download { display: inline-block; margin-top: 20px; margin-right: 10px; padding: 10px 20px; background-color: #28a745; color: white; text-decoration: none; }
    .download:hover { background-color: #218838; }
    .download-excel { background-color: #17a2b8; }
    .download-excel:hover { background-color: #138496; }
    #chart { margin-top: 20px; }
</style>
"""

# Upload page HTML
upload_html = """
<!DOCTYPE html>
<html>
<head><title>Price Deconstructor</title>
""" + css + """
</head>
<body>
    <div class="container">
        <h1>Upload Sales Report</h1>
        {{error|safe}}
        <form method="post" enctype="multipart/form-data">
            <div class="form-group">
                <label for="file">Select Excel File (.xlsx):</label>
                <input type="file" id="file" name="file" accept=".xlsx">
            </div>
            <button type="submit">Upload & Proceed to Pricing</button>
        </form>
        <p><a href="/debug">View Debug Info</a></p>
    </div>
</body>
</html>
"""

# Pricing form HTML
pricing_form_html = """
<!DOCTYPE html>
<html>
<head><title>Pricing Rules</title>
""" + css + """
</head>
<body>
    <div class="container">
        <h1>Enter Pricing Rules</h1>
        {{error|safe}}
        <form method="post" action="/pricing" enctype="multipart/form-data">
            <div class="form-group">
                <label for="pricing_file">Import Pricing File (.txt):</label>
                <input type="file" id="pricing_file" name="pricing_file" accept=".txt">
            </div>
            <h3>Process and Step Process</h3>
            {% for process in processes %}
            <div class="form-group">
                <h4>{{process}}</h4>
                {% for step in process_step_mapping[process] %}
                <div class="form-group">
                    <label for="{{process}}_{{step}}">{{step}}</label>
                    <input type="number" step="0.01" id="{{process}}_{{step}}" name="{{process}}_{{step}}" placeholder="Cost ($)" value="{{form_data.get(process ~ '_' ~ step, '')}}">
                </div>
                {% endfor %}
            </div>
            {% endfor %}
            <h3>Coating</h3>
            {% for coating in ['Advanced Nano', 'Nano Wipe', 'Nano Slic', 'BluPrint'] %}
            <div class="form-group">
                <label for="Coating_{{coating}}">{{coating}}</label>
                <input type="number" step="0.01" id="Coating_{{coating}}" name="Coating_{{coating}}" placeholder="Cost ($)" value="{{form_data.get('Coating_' ~ coating, '')}}">
            </div>
            {% endfor %}
            <button type="submit">Process File</button>
        </form>
        <p><a href="/debug">View Debug Info</a></p>
    </div>
</body>
</html>
"""

# Results page HTML
results_html = """
<!DOCTYPE html>
<html>
<head>
    <title>Results</title>
""" + css + """
    <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
</head>
<body>
    <div class="container">
        <h1>Deconstructed Pricing</h1>
        <p>{{data|length}} Unique Customer-Material-Price Combinations Processed</p>
        {{error|safe}}
        <a href="/download" class="download">Download Results as CSV</a>
        <a href="/download_excel" class="download download-excel">Download Results as Excel</a>
        <h3>Lowest Base Cost by Customer</h3>
        <div id="chart">{{chart|safe}}</div>
        <table>
            <thead>
                <tr>
                    <th>Customer</th>
                    <th>Customer Internal ID</th>
                    <th>Frame</th>
                    <th>Item Internal ID</th>
                    <th>Sales Price</th>
                    <th>Process</th>
                    <th>Step Process</th>
                    <th>Coating</th>
                    <th>Foil Material</th>
                    <th>Foil Thickness</th>
                    <th>Colour</th>
                    <th>Attribute Cost</th>
                    <th>Base Cost</th>
                </tr>
            </thead>
            <tbody>
                {% for row in data %}
                <tr>
                    <td>{{row.Customer}}</td>
                    <td>{{row.Customer_Internal_ID}}</td>
                    <td>{{row.Frame}}</td>
                    <td>{{row.Item_Internal_ID}}</td>
                    <td>{{row.Sales_Price}}</td>
                    <td>{{row.Process}}</td>
                    <td>{{row.Step_Process}}</td>
                    <td>{{row.Coating}}</td>
                    <td>{{row.Foil_Material}}</td>
                    <td>{{row.Foil_Thickness}}</td>
                    <td>{{row.Colour}}</td>
                    <td>{{row.Attribute_Cost}}</td>
                    <td>{{row.Base_Cost}}</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
        <p><a href="/debug">View Debug Info</a></p>
    </div>
</body>
</html>
"""

# Debug page HTML
debug_html = """
<!DOCTYPE html>
<html>
<head><title>Debug Info</title>
""" + css + """
</head>
<body>
    <div class="container">
        <h1>Debug Information</h1>
        <p class="debug">Timestamp: {{timestamp}}</p>
        <p class="debug">Session File Path: {{file_path}}</p>
        <p class="debug">File Exists: {{file_exists}}</p>
        <p class="debug">Uploads Folder Contents: {{uploads_contents}}</p>
        <p class="debug">Sheet Names: {{sheet_names}}</p>
        <p class="debug">Column Names: {{column_names}}</p>
        <p class="debug">Form Data: {{form_data}}</p>
        <p class="debug">Session Data: {{session_data}}</p>
        <p><a href="/">Back to Upload</a></p>
    </div>
</body>
</html>
"""

@app.route('/debug')
def debug_info():
    file_path = session.get('file_path', 'None')
    file_exists = os.path.exists(file_path) if file_path != 'None' else False
    uploads_contents = os.listdir(UPLOAD_FOLDER)
    sheet_names = 'None'
    column_names = 'None'
    form_data = session.get('form_data', 'None')
    session_data = dict(session)  # Get all session data for debugging
    if file_exists:
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet_names = ', '.join(wb.sheetnames)
            logger.debug(f"Sheet names in {file_path}: {sheet_names}")
            if 'SalesbyItemBASEPRICEDECON' in wb.sheetnames:
                df = pd.read_excel(file_path, sheet_name='SalesbyItemBASEPRICEDECON', engine='openpyxl', nrows=1)
                column_names = ', '.join(str(col) for col in df.columns)
                logger.debug(f"Column names in {file_path}: {column_names}")
            else:
                column_names = 'Sheet not found'
        except Exception as e:
            sheet_names = f'Error reading sheets: {str(e)}'
            column_names = 'N/A'
            logger.error(f"Error reading sheet names or columns: {str(e)}")
    return app.jinja_env.from_string(debug_html).render(
        timestamp=datetime.now().strftime('%Y%m%d_%H%M%S'),
        file_path=file_path,
        file_exists=file_exists,
        uploads_contents=', '.join(uploads_contents) if uploads_contents else 'Empty',
        sheet_names=sheet_names,
        column_names=column_names,
        form_data=form_data,
        session_data=session_data
    )

def sanitize_filename(filename):
    """Sanitize filename by removing or replacing problematic characters."""
    valid_chars = "-_.() %s%s" % (string.ascii_letters, string.digits)
    sanitized = ''.join(c if c in valid_chars else '_' for c in filename)
    sanitized = re.sub(r'_+', '_', sanitized)
    return sanitized.strip('_')

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    logger.debug("Entering / route")
    if request.method == 'POST':
        logger.debug("Received POST request for file upload")
        try:
            file = request.files.get('file')
            if not file:
                logger.error("No file provided in upload")
                return app.jinja_env.from_string(upload_html).render(error='<p class="error">No file selected. Please choose a file.</p>')
            
            if not file.filename.endswith('.xlsx'):
                logger.error(f"Invalid file extension: {file.filename}")
                return app.jinja_env.from_string(upload_html).render(error='<p class="error">Please upload a valid .xlsx file.</p>')
            
            # Generate a unique filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            sanitized_filename = sanitize_filename(file.filename)
            file_path = os.path.join(UPLOAD_FOLDER, f"{timestamp}_{sanitized_filename}")
            file_path = os.path.normpath(file_path)  # Normalize path for consistency
            logger.debug(f"Saving file to: {file_path}")
            
            # Check write permissions for the Uploads folder
            if not os.access(UPLOAD_FOLDER, os.W_OK):
                logger.error(f"No write permissions for Uploads folder: {UPLOAD_FOLDER}")
                return app.jinja_env.from_string(upload_html).render(error='<p class="error">Server error: No write permissions for Uploads folder. Please contact the administrator.</p>')
            
            # Save the file
            file.save(file_path)
            
            # Verify file exists after saving
            if not os.path.exists(file_path):
                logger.error(f"File not found after saving: {file_path}")
                return app.jinja_env.from_string(upload_html).render(error=f'<p class="error">Failed to save file: {file.filename}. Please check disk space or permissions and try again.</p>')
            
            # Log file permissions
            file_stats = os.stat(file_path)
            logger.debug(f"File permissions for {file_path}: {oct(file_stats.st_mode)[-3:]}")
            
            # Validate file structure
            logger.debug(f"Validating Excel file structure: {file_path}")
            wb = openpyxl.load_workbook(file_path)
            sheet_names = wb.sheetnames
            logger.debug(f"Sheet names: {sheet_names}")
            if 'SalesbyItemBASEPRICEDECON' not in sheet_names:
                logger.error(f"Sheet 'SalesbyItemBASEPRICEDECON' not found in {file_path}")
                try:
                    os.remove(file_path)
                    logger.debug(f"Removed invalid file: {file_path}")
                except Exception as e:
                    logger.warning(f"Failed to remove invalid file {file_path}: {str(e)}")
                return app.jinja_env.from_string(upload_html).render(error=f'<p class="error">Sheet "SalesbyItemBASEPRICEDECON" not found in {file.filename}. Available sheets: {", ".join(sheet_names)}</p>')
            
            df = pd.read_excel(file_path, sheet_name='SalesbyItemBASEPRICEDECON', engine='openpyxl', nrows=1)
            actual_columns = [str(col).strip().lower() for col in df.columns]
            logger.debug(f"Actual columns: {', '.join(df.columns)}")
            required_columns = [
                'Sales Price', 'Frame', 'Customer/Project: Company Name',
                'Process', '[ES] Step Process', 'Coating', 'Foil Material',
                'Foil Thickness', 'Colour'
            ]
            optional_columns = ['Customer/Project: Internal ID', 'Item: Internal ID']
            required_columns_normalized = [col.strip().lower() for col in required_columns]
            missing_required_columns = [col for col in required_columns if col.strip().lower() not in actual_columns]
            missing_optional_columns = [col for col in optional_columns if col.strip().lower() not in actual_columns]
            if missing_required_columns:
                logger.warning(f"Missing required columns in Excel file: {missing_required_columns}. Cannot proceed.")
                try:
                    os.remove(file_path)
                    logger.debug(f"Removed invalid file: {file_path}")
                except Exception as e:
                    logger.warning(f"Failed to remove invalid file {file_path}: {str(e)}")
                return app.jinja_env.from_string(upload_html).render(error=f'<p class="error">Missing required columns in {file.filename}: {", ".join(missing_required_columns)}. Found: {", ".join(df.columns)}</p>')
            if missing_optional_columns:
                logger.warning(f"Missing optional columns in Excel file: {missing_optional_columns}. Proceeding with warning.")
                session['column_warning'] = f"Missing optional columns in {file.filename}: {', '.join(missing_optional_columns)}. Found: {', '.join(df.columns)}"
            else:
                session['column_warning'] = None
                logger.debug("Excel file validated successfully")
            
            # Store file path in session and make it permanent
            session.permanent = True  # Persist session for the configured lifetime
            session['file_path'] = file_path
            session['form_data'] = 'None'  # Reset form data
            logger.debug(f"File uploaded and saved: {file_path}, stored in session['file_path']")
            
            logger.debug("Rendering pricing form after successful upload")
            return app.jinja_env.from_string(pricing_form_html).render(
                processes=process_step_mapping.keys(),
                process_step_mapping=process_step_mapping,
                form_data={},
                error=None
            )
        except Exception as e:
            logger.error(f"Unexpected error during file upload/validation: {str(e)}")
            return app.jinja_env.from_string(upload_html).render(error=f'<p class="error">Unexpected error during upload of {file.filename if file else "file"}: {str(e)}. Please ensure the file is accessible and try again.</p>')
    
    logger.debug("Rendering upload page for GET request")
    return app.jinja_env.from_string(upload_html).render(error=None)

@app.route('/pricing', methods=['GET', 'POST'])
def pricing_form():
    if request.method == 'GET':
        logger.debug("Accessed /pricing via GET, redirecting to upload page")
        return app.jinja_env.from_string(upload_html).render(error='<p class="error">Please upload a file first.</p>')
    
    logger.debug("Processing pricing form submission")
    
    global pricing_rules
    pricing_rules = {
        "Process": {},
        "Coating": {},
        "Foil Material": {},
        "Foil Thickness": {},
        "Colour": {}
    }
    
    # Initialize form_data
    form_data = {}
    
    # Check if a pricing file was uploaded
    pricing_file = request.files.get('pricing_file')
    if pricing_file and pricing_file.filename.endswith('.txt'):
        try:
            logger.debug(f"Processing uploaded pricing file: {pricing_file.filename}")
            # Save the pricing file temporarily
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            sanitized_pricing_filename = sanitize_filename(pricing_file.filename)
            pricing_file_path = os.path.join(UPLOAD_FOLDER, f"{timestamp}_{sanitized_pricing_filename}")
            pricing_file_path = os.path.normpath(pricing_file_path)
            logger.debug(f"Saving pricing file to: {pricing_file_path}")
            pricing_file.save(pricing_file_path)
            
            # Verify pricing file exists
            if not os.path.exists(pricing_file_path):
                logger.error(f"Pricing file not found after saving: {pricing_file_path}")
                return app.jinja_env.from_string(pricing_form_html).render(
                    processes=process_step_mapping.keys(),
                    process_step_mapping=process_step_mapping,
                    form_data=form_data,
                    error=f'<p class="error">Failed to save pricing file: {pricing_file.filename}. Please check disk space or permissions and try again.</p>'
                )
            
            # Parse the pricing file
            with open(pricing_file_path, 'r') as f:
                for line in f:
                    line = line.strip()
                    if ':' not in line:
                        continue
                    key, value = [part.strip() for part in line.split(':', 1)]
                    try:
                        value = float(value)
                    except ValueError:
                        logger.warning(f"Invalid price value in pricing file for {key}: {value}")
                        continue
                    
                    # Normalize key for comparison
                    key = key.lower().replace('lasterstep', 'laserstep').replace('laststep', 'laserstep')
                    
                    # Map keys to form_data
                    if key.startswith('chem '):
                        step = key[5:]  # Keep exact format (e.g., "5 or more")
                        if step == '5 or more' or step.title() in process_step_mapping["Chemetch"]:
                            form_data[f"Chemetch_{step if step == '5 or more' else step.title()}"] = str(value)
                            logger.debug(f"Set form_data[Chemetch_{step if step == '5 or more' else step.title()}]: {value}")
                    elif key.startswith('laserstep '):
                        step = key[10:]  # Keep exact format (e.g., "1-2")
                        if step in process_step_mapping["LaserSTEP"]:
                            form_data[f"LaserSTEP_{step}"] = str(value)
                            logger.debug(f"Set form_data[LaserSTEP_{step}]: {value}")
                        # Default to 1-20 price for new ranges if not specified
                        elif step in ["21-30", "31-40", "41-50", "51-60"]:
                            form_data[f"LaserSTEP_{step}"] = str(245)  # Use 1-20 price
                            logger.debug(f"Set form_data[LaserSTEP_{step}]: 245 (default from 1-20)")
                    elif key.startswith('mill '):
                        step = key[5:].title()  # Convert to title case (e.g., "single" → "Single")
                        if step in process_step_mapping["Milled"]:
                            form_data[f"Milled_{step}"] = str(value)
                            logger.debug(f"Milled_{step}: {value}")
                    elif key == 'double':  # Handle ambiguous "double" (assume Milled_Double)
                        form_data["Milled_Double"] = str(value)
                        logger.warning(f"Ambiguous key 'double' mapped to Milled_Double: {value}")
                    elif key.startswith('coat '):
                        coating = key[5:].title().replace('Bluprint', 'BluPrint')  # Handle title case and BluPrint
                        if coating in ["Advanced Nano", "Nano Wipe", "Nano Slic", "BluPrint"]:
                            form_data[f"Coating_{coating}"] = str(value)
                            logger.debug(f"Set form_data[Coating_{coating}]: {value}")
            
            # Clean up the pricing file
            try:
                os.remove(pricing_file_path)
                logger.debug(f"Removed temporary pricing file: {pricing_file_path}")
            except Exception as e:
                logger.warning(f"Failed to remove pricing file {pricing_file_path}: {str(e)}")
            
            # Log form_data for debugging
            logger.debug(f"Form data after pricing file parsing: {form_data}")
            
            # Ensure session is still valid
            if not session.get('file_path'):
                logger.error("Session file_path missing after pricing file upload")
                return app.jinja_env.from_string(upload_html).render(error='<p class="error">Session expired or no Excel file uploaded. Please upload the Excel file again.</p>')
            
            # Render the pricing form with pre-filled values
            return app.jinja_env.from_string(pricing_form_html).render(
                processes=process_step_mapping.keys(),
                process_step_mapping=process_step_mapping,
                form_data=form_data,
                error=None
            )
        except Exception as e:
            logger.error(f"Error processing pricing file: {str(e)}")
            return app.jinja_env.from_string(pricing_form_html).render(
                processes=process_step_mapping.keys(),
                process_step_mapping=process_step_mapping,
                form_data=form_data,
                error=f'<p class="error">Error processing pricing file: {pricing_file.filename}. Please check disk space or permissions and try again.</p>'
            )
    
    # Process form data (manual entry or after file import)
    try:
        form_data = {key: value for key, value in request.form.items()}
        session['form_data'] = str(form_data)[:1000]  # Truncate for debug display
        logger.debug(f"Form data received: {form_data}")
        non_zero_prices = False
        for process in process_step_mapping:
            pricing_rules["Process"][process] = {}
            for step in process_step_mapping[process]:
                cost = request.form.get(f"{process}_{step}", "0")
                try:
                    cost_value = float(cost) if cost.strip() else 0
                    # Apply 1-20 price (245) for LaserSTEP ranges >= 21-30 if not specified
                    if process == "LaserSTEP" and step in ["21-30", "31-40", "41-50", "51-60"] and cost_value == 0:
                        cost_value = pricing_rules["Process"]["LaserSTEP"].get("1-20", 245)
                        logger.debug(f"Applied default price for {process}_{step}: {cost_value} (from 1-20)")
                    pricing_rules["Process"][process][step] = cost_value
                    if cost_value != 0:
                        non_zero_prices = True
                    logger.debug(f"Set price for {process}_{step}: {cost_value}")
                except ValueError:
                    logger.warning(f"Invalid cost value for {process}_{step}: {cost}")
                    pricing_rules["Process"][process][step] = 0
        
        for coating in ["Advanced Nano", "Nano Wipe", "Nano Slic", "BluPrint"]:
            cost = request.form.get(f"Coating_{coating}", "0")
            try:
                cost_value = float(cost) if cost.strip() else 0
                pricing_rules["Coating"][coating] = cost_value
                if cost_value != 0:
                    non_zero_prices = True
                logger.debug(f"Set price for Coating_{coating}: {cost_value}")
            except ValueError:
                logger.warning(f"Invalid cost value for Coating_{coating}: {cost}")
                pricing_rules["Coating"][coating] = 0
        
        # Log pricing_rules for debugging
        logger.debug(f"Final pricing_rules: {pricing_rules}")
        
        # Check if any non-zero prices were set
        if not non_zero_prices:
            logger.warning("No non-zero pricing rules provided")
            return app.jinja_env.from_string(pricing_form_html).render(
                processes=process_step_mapping.keys(),
                process_step_mapping=process_step_mapping,
                form_data=form_data,
                error='<p class="error">Please provide at least one non-zero pricing rule.</p>'
            )
    except Exception as e:
        logger.error(f"Error processing form data: {str(e)}")
        return app.jinja_env.from_string(pricing_form_html).render(
            processes=process_step_mapping.keys(),
            process_step_mapping=process_step_mapping,
            form_data=form_data,
            error=f'<p class="error">Error processing pricing form: {str(e)}. Please try again.</p>'
        )
    
    # Process uploaded Excel file
    file_path = session.get('file_path')
    logger.debug(f"Checking session file_path: {file_path}")
    if not file_path:
        logger.error("No file path found in session")
        return app.jinja_env.from_string(upload_html).render(error='<p class="error">No Excel file path found in session. Please upload the Excel file again. Ensure cookies are enabled in your browser.</p>')
    if not os.path.exists(file_path):
        logger.error(f"File does not exist on disk: {file_path}")
        return app.jinja_env.from_string(upload_html).render(error=f'<p class="error">Uploaded Excel file not found on disk: {os.path.basename(file_path)}. It may have been deleted, moved, or not saved properly. Please upload again.</p>')
    
    try:
        logger.debug(f"Validating file before processing: {file_path}")
        # Check file permissions
        if not os.access(file_path, os.R_OK):
            logger.error(f"No read permissions for file: {file_path}")
            try:
                os.remove(file_path)
                logger.debug(f"Removed invalid file: {file_path}")
            except Exception as e:
                logger.warning(f"Failed to remove invalid file {file_path}: {str(e)}")
            return app.jinja_env.from_string(upload_html).render(error=f'<p class="error">No read permissions for file: {os.path.basename(file_path)}. Please check file permissions and upload again.</p>')
        
        wb = openpyxl.load_workbook(file_path)
        sheet_names = wb.sheetnames
        logger.debug(f"Sheet names: {sheet_names}")
        if 'SalesbyItemBASEPRICEDECON' not in sheet_names:
            logger.error(f"Sheet 'SalesbyItemBASEPRICEDECON' not found in {file_path}")
            try:
                os.remove(file_path)
                logger.debug(f"Removed invalid file: {file_path}")
            except Exception as e:
                logger.warning(f"Failed to remove invalid file {file_path}: {str(e)}")
            return app.jinja_env.from_string(upload_html).render(error=f'<p class="error">Sheet "SalesbyItemBASEPRICEDECON" not found in {os.path.basename(file_path)}. Available sheets: {", ".join(sheet_names)}</p>')
        
        df = pd.read_excel(file_path, sheet_name='SalesbyItemBASEPRICEDECON', engine='openpyxl')
        logger.debug(f"Excel file read successfully: {file_path}, {len(df)} rows")
        actual_columns = [str(col).strip().lower() for col in df.columns]
        logger.debug(f"Actual columns: {', '.join(df.columns)}")
        required_columns = [
            'Sales Price', 'Frame', 'Customer/Project: Company Name',
            'Process', '[ES] Step Process', 'Coating', 'Foil Material',
            'Foil Thickness', 'Colour'
        ]
        optional_columns = ['Customer/Project: Internal ID', 'Item: Internal ID']
        required_columns_normalized = [col.strip().lower() for col in required_columns]
        missing_required_columns = [col for col in required_columns if col.strip().lower() not in actual_columns]
        missing_optional_columns = [col for col in optional_columns if col.strip().lower() not in actual_columns]
        if missing_required_columns:
            logger.error(f"Missing required columns in Excel file: {missing_required_columns}")
            try:
                os.remove(file_path)
                logger.debug(f"Removed invalid file: {file_path}")
            except Exception as e:
                logger.warning(f"Failed to remove invalid file {file_path}: {str(e)}")
            return app.jinja_env.from_string(upload_html).render(error=f'<p class="error">Missing required columns in {os.path.basename(file_path)}: {", ".join(missing_required_columns)}. Found: {", ".join(df.columns)}</p>')
        if missing_optional_columns:
            logger.warning(f"Missing optional columns in Excel file: {missing_optional_columns}. Proceeding with warning.")
            session['column_warning'] = f"Missing optional columns in {os.path.basename(file_path)}: {', '.join(missing_optional_columns)}. Found: {', '.join(df.columns)}"
        else:
            session['column_warning'] = None
            logger.debug("Excel file validated successfully")
        
        results = []
        skipped_rows = []
        for index, row in df.iterrows():
            try:
                # Check for missing required fields
                missing_fields = []
                if pd.isna(row['Sales Price']):
                    missing_fields.append('Sales Price')
                if pd.isna(row['Frame']):
                    missing_fields.append('Frame')
                if pd.isna(row['Customer/Project: Company Name']):
                    missing_fields.append('Customer/Project: Company Name')
                if missing_fields:
                    skipped_rows.append((index, f"Missing required fields: {', '.join(missing_fields)}"))
                    logger.debug(f"Skipping row {index} due to missing required fields: {', '.join(missing_fields)}")
                    continue
                
                process = str(row['Process']).strip() if not pd.isna(row['Process']) else 'Unknown'
                step_process = str(row['[ES] Step Process']).strip() if not pd.isna(row['[ES] Step Process']) else 'None'
                if process == 'LaserSTEP':
                    step_process = re.sub(r'\s*-\s*', '-', step_process)
                    logger.debug(f"Normalized step_process for LaserSTEP: {step_process}")
                coating = str(row['Coating']).strip() if not pd.isna(row['Coating']) else 'None'
                foil_material = str(row['Foil Material']).strip() if not pd.isna(row['Foil Material']) else 'Unknown'
                foil_thickness = str(row['Foil Thickness']).strip() if not pd.isna(row['Foil Thickness']) else 'Unknown'
                colour = str(row['Colour']).strip() if not pd.isna(row['Colour']) else 'Unknown'
                customer = str(row['Customer/Project: Company Name']).strip() if not pd.isna(row['Customer/Project: Company Name']) else 'Unknown'
                customer_internal_id = str(row.get('Customer/Project: Internal ID', 'Unknown')).strip()
                item_internal_id = str(row.get('Item: Internal ID', 'Unknown')).strip()
                
                try:
                    sales_price = float(row['Sales Price'])
                except (ValueError, TypeError) as e:
                    skipped_rows.append((index, f"Invalid Sales Price: {row['Sales Price']}"))
                    logger.warning(f"Invalid Sales Price in row {index}: {row['Sales Price']}")
                    continue
                
                attribute_cost = 0
                if process != 'LaserCut':
                    if process in pricing_rules["Process"]:
                        logger.debug(f"Available steps for {process}: {list(pricing_rules['Process'][process].keys())}")
                        if step_process in pricing_rules["Process"][process]:
                            attribute_cost += pricing_rules["Process"][process][step_process]
                            logger.debug(f"Applied process cost: {process}_{step_process} = {pricing_rules['Process'][process][step_process]}")
                        else:
                            logger.warning(f"Invalid step_process in row {index}: {step_process} for process {process}")
                    else:
                        logger.warning(f"Invalid process in row {index}: {process}")
                    if coating in pricing_rules["Coating"]:
                        attribute_cost += pricing_rules["Coating"][coating]
                        logger.debug(f"Applied coating cost: Coating_{coating} = {pricing_rules['Coating'][coating]}")
                    else:
                        logger.warning(f"Invalid coating in row {index}: {coating}")
                
                base_cost = sales_price - attribute_cost
                
                results.append({
                    'Customer': customer,
                    'Customer_Internal_ID': customer_internal_id,
                    'Frame': str(row['Frame']).strip(),
                    'Item_Internal_ID': item_internal_id,
                    'Sales_Price': sales_price,
                    'Process': process,
                    'Step_Process': step_process,
                    'Coating': coating,
                    'Foil_Material': foil_material,
                    'Foil_Thickness': foil_thickness,
                    'Colour': colour,
                    'Attribute_Cost': attribute_cost,
                    'Base_Cost': base_cost
                })
            except Exception as e:
                skipped_rows.append((index, f"Error processing row: {str(e)}"))
                logger.warning(f"Error processing row {index}: {str(e)}")
                continue
        
        if not results:
            logger.error(f"No valid data processed from Excel file. Skipped {len(skipped_rows)} rows.")
            error_message = f'<p class="error">No valid data found in Excel file {os.path.basename(file_path)}. Reasons for skipping rows:<br>'
            error_message += '<ul>' + ''.join(f'<li>Row {row_idx}: {reason}</li>' for row_idx, reason in skipped_rows[:10]) + '</ul>'
            if len(skipped_rows) > 10:
                error_message += f'<p>And {len(skipped_rows) - 10} more rows skipped. Check the debug log for details.</p>'
            error_message += '<p>Please check the file contents (e.g., ensure Sales Price, Frame, and Customer/Project: Company Name are populated).</p>'
            try:
                os.remove(file_path)
                logger.debug(f"Removed invalid file: {file_path}")
            except Exception as e:
                logger.warning(f"Failed to remove invalid file {file_path}: {str(e)}")
            return app.jinja_env.from_string(upload_html).render(error=error_message)
        
        # Remove duplicates by customer, material, and sales price combination
        try:
            result_df = pd.DataFrame(results)
            logger.debug(f"Processed {len(result_df)} rows before duplicate removal")
            if result_df.empty:
                logger.error("DataFrame is empty after processing")
                try:
                    os.remove(file_path)
                    logger.debug(f"Removed invalid file: {file_path}")
                except Exception as e:
                    logger.warning(f"Failed to remove invalid file {file_path}: {str(e)}")
                return app.jinja_env.from_string(upload_html).render(error=f'<p class="error">No valid data after processing {os.path.basename(file_path)}. Please check the file contents.</p>')
            # Ensure Customer and Sales_Price are valid
            if 'Customer' not in result_df.columns or 'Sales_Price' not in result_df.columns:
                logger.error(f"Missing critical columns in DataFrame: {result_df.columns}")
                try:
                    os.remove(file_path)
                    logger.debug(f"Removed invalid file: {file_path}")
                except Exception as e:
                    logger.warning(f"Failed to remove invalid file {file_path}: {str(e)}")
                return app.jinja_env.from_string(upload_html).render(error=f'<p class="error">Missing critical columns in {os.path.basename(file_path)}: {", ".join(result_df.columns)}</p>')
            # Handle non-string Customers or non-numeric Sales_Price
            result_df['Customer'] = result_df['Customer'].astype(str)
            result_df['Customer_Internal_ID'] = result_df['Customer_Internal_ID'].astype(str)
            result_df['Item_Internal_ID'] = result_df['Item_Internal_ID'].astype(str)
            result_df['Sales_Price'] = pd.to_numeric(result_df['Sales_Price'], errors='coerce')
            result_df['Base_Cost'] = pd.to_numeric(result_df['Base_Cost'], errors='coerce')
            if result_df['Sales_Price'].isna().all():
                logger.error("All Sales_Price values are invalid")
                try:
                    os.remove(file_path)
                    logger.debug(f"Removed invalid file: {file_path}")
                except Exception as e:
                    logger.warning(f"Failed to remove invalid file {file_path}: {str(e)}")
                return app.jinja_env.from_string(upload_html).render(error=f'<p class="error">All Sales_Price values are invalid in {os.path.basename(file_path)}. Please check Sales Price data.</p>')
            # Define columns for deduplication
            dedup_columns = ['Customer', 'Process', 'Step_Process', 'Coating', 'Foil_Material', 'Foil_Thickness', 'Colour', 'Sales_Price']
            # Remove exact duplicates based on customer, material attributes, and sales price
            result_df = result_df.drop_duplicates(subset=dedup_columns, keep='first').reset_index(drop=True)
            logger.debug(f"After duplicate removal: {len(result_df)} unique customer-material-price combinations")
        except Exception as e:
            logger.error(f"Error processing results: {str(e)}")
            try:
                os.remove(file_path)
                logger.debug(f"Removed invalid file: {file_path}")
            except Exception as e:
                logger.warning(f"Failed to remove invalid file {file_path}: {str(e)}")
            return app.jinja_env.from_string(upload_html).render(error=f'<p class="error">Error processing results from {os.path.basename(file_path)}: {str(e)}. Please try again.</p>')
        
        # Generate bar chart for lowest Base Cost by Customer
        try:
            if result_df.empty:
                logger.error("DataFrame is empty for chart generation")
                chart_html = '<p class="error">No data available for chart</p>'
            else:
                # For the chart, group by Customer and take the minimum Base_Cost
                chart_df = result_df.loc[result_df.groupby('Customer')['Base_Cost'].idxmin()]
                fig = px.bar(chart_df, x='Customer', y='Base_Cost', title='Lowest Base Cost by Customer',
                             labels={'Base_Cost': 'Base Cost ($)', 'Customer': 'Customer'})
                fig.update_layout(xaxis_tickangle=45)
                chart_html = pio.to_html(fig, full_html=False)
                logger.debug("Bar chart generated successfully")
        except Exception as e:
            logger.error(f"Error generating chart: {str(e)}")
            chart_html = f'<p class="error">Error generating chart: {str(e)}</p>'
        
        # Save results to CSV and Excel
        csv_path = os.path.join(UPLOAD_FOLDER, 'results.csv')
        excel_path = os.path.join(UPLOAD_FOLDER, 'results.xlsx')
        try:
            result_df.to_csv(csv_path, index=False)
            result_df.to_excel(excel_path, index=False, engine='openpyxl')
            logger.debug(f"Results saved to {csv_path} and {excel_path}")
        except Exception as e:
            logger.error(f"Error saving results: {str(e)}")
            try:
                os.remove(file_path)
                logger.debug(f"Removed invalid file: {file_path}")
            except Exception as e:
                logger.warning(f"Failed to remove invalid file {file_path}: {str(e)}")
            return app.jinja_env.from_string(upload_html).render(error=f'<p class="error">Error saving results: {str(e)}. Please try again.</p>')
        
        # Clean up the uploaded Excel file after processing
        try:
            os.remove(file_path)
            logger.debug(f"Removed uploaded Excel file: {file_path}")
        except Exception as e:
            logger.warning(f"Failed to remove Excel file {file_path}: {str(e)}")
        
        # Include column warning if any
        column_warning = session.get('column_warning')
        if column_warning:
            logger.debug(f"Rendering results with column warning: {column_warning}")
            return app.jinja_env.from_string(results_html).render(
                data=result_df.to_dict('records'),
                chart=chart_html,
                error=f'<p class="error">Warning: {column_warning}</p>'
            )
        
        logger.debug("Rendering results page")
        return app.jinja_env.from_string(results_html).render(data=result_df.to_dict('records'), chart=chart_html)
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        try:
            os.remove(file_path)
            logger.debug(f"Removed invalid file: {file_path}")
        except Exception as e:
            logger.warning(f"Failed to remove invalid file {file_path}: {str(e)}")
        return app.jinja_env.from_string(upload_html).render(error=f'<p class="error">Error reading Excel file {os.path.basename(file_path)}: {str(e)}. Please upload again.</p>')

@app.route('/download')
def download_csv():
    result_path = os.path.join(UPLOAD_FOLDER, 'results.csv')
    if os.path.exists(result_path):
        response = make_response(open(result_path, 'rb').read())
        response.headers['Content-Disposition'] = 'attachment; filename=results.csv'
        response.mimetype = 'text/csv'
        logger.debug(f"Serving CSV download: {result_path}")
        return response
    logger.error("CSV file not found for download")
    return app.jinja_env.from_string(upload_html).render(error='<p class="error">No results available for download. Please process the file again.</p>')

@app.route('/download_excel')
def download_excel():
    result_path = os.path.join(UPLOAD_FOLDER, 'results.xlsx')
    if os.path.exists(result_path):
        response = make_response(open(result_path, 'rb').read())
        response.headers['Content-Disposition'] = 'attachment; filename=results.xlsx'
        response.mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        logger.debug(f"Serving Excel download: {result_path}")
        return response
    logger.error("Excel file not found for download")
    return app.jinja_env.from_string(upload_html).render(error='<p class="error">No results available for download. Please process the file again.</p>')

if __name__ == '__main__':
    app.run(debug=True)